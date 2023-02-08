from datetime import datetime
from openpyxl import load_workbook


Rut="C:\Users\SENA\Documents\jasly\danilo python\APPCRUD.PY\BaseCrud.xlsx"
Rut=r"C:\Users\SENA\Documents\jasly\danilo python\APPCRUD.PY\BaseCrud.xlsx"


def leer (ruta:str, extraer:str):
    archivo_exccel = load_workbook(ruta)
    hoja_datos=archivo_exccel['datos del crud']
    hoja_datos=hoja_datos['A2':'f'+str(hoja_datos.max_row)]

    info={}


    for i in hoja_datos:

        if isinstance(i[0].value,int):
            info.setdefault(i[0].value,{'tarea':i[1].value,'descripcion':i[2].value,
                                        'estado':i[3].value,'fecha de inicio':i[4].value,
                                        'fecha de finalizacion':i[5].value})


    if not (extraer=='todo'):
        info=filtrar(info, extraer)


    for i in info:
        print ('******** tarea *******')
        print('Id'+ str(i)+'\n'+'titulo:'+str(info[i]['tarea'])+'\n'+'descripcion:'
        +str(info[i]['descripcion'])+'\n'+'estado:'+str(info[i]['estado'])
        +'\n'+'fecha creacion:'+str(info[i]['fecha de inicio'])
        +'\n'+'fecha de finalizacion:'+str(info[i]['fecha de finalizacion']))
        print()

    return    



def filtrar (info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux



def actualizar(ruta:str,identificador:int , datos_actualizados:dict):
    archivo_excel= load_workbook(ruta)
    hoja_datos=archivo_excel['datos del crud']
    hoja_datos=hoja_datos['A2':'f'+str(hoja_datos.max_row)]
    hoja=archivo_excel.active


    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d== 'titulo 'and not (datos_actualizados[d]==''):
                    hoja.active_cell(row=fila,column=titulo).value=datos_actualizados[d]
                elif d=='descripcion' and not (datos_actualizados[d]==''):
                    hoja.active_cell(row=fila, column =descripcion).value =datos_actualizados[d]  
                elif d== 'estado'and not (datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=estado).value=datos_actualizados[d]
                elif d== 'fecha inicio' and not (datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_inicio).value=datos_actualizados[d]
                elif d=='fecha finalizacion' and not (datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=fecha_finalizado).value=datos_actualizados[d]
    archivo_excel.save(ruta)
    if encontro==False:
        print('error: No exist euna tarea con ese Id')
        print()
    return


def agregar(ruta:int, datos:dict):
    archivo_exccel = load_workbook(ruta)
    hoja_datos = archivo_exccel['datos del crud']
    hoja_datos=hoja_datos['A2':'F'+str(hoja_datos.max_row+1)]
    hoja=archivo_exccel.active



    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False
    for i in hoja_datos:

        if not (isinstance(i[0].value,int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador, column=fecha_finalizado).value=datos['fecha finalizado']
            break
    archivo_exccel.save(ruta)
    return



def borrar(ruta,identificador):
    archivo_exccel=load_workbook(ruta)
    hoja_datos=archivo_exccel['datos del crud']
    hoja_datos=hoja_datos['A2':'F'+str(hoja_datos.max_row)]
    hoja=archivo_exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True


        





        
       





